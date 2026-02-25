#!/usr/bin/env python3
"""
Compare the SAME source file across two run log xlsx files.
Each row = one document (source file); columns = OCR/Summary from run1, OCR/Summary from run2,
then for both OCR and Invoice summary: Difference (run1 vs run2), Same % (character accuracy),
plus Metadata consistent (Gemini) and Summary of what was inconsistent.

- OCR: Difference (unified diff), Same % (OCR) – character-level comparison.
- Summary: Difference (unified diff), Same % (Summary) – character-level comparison.
- Metadata consistent / Summary of what was inconsistent: Gemini logical comparison of the two summaries.

Usage:
  python compare_ocr_run_log.py file1.xlsx file2.xlsx
  python compare_ocr_run_log.py log_files/run1.xlsx log_files/run2.xlsx --output comparison.xlsx
  python compare_ocr_run_log.py --no-gemini   # skip LLM column

Exactly 2 run log files must be passed. Default output: log_files/ocr_comparison_YYYY-MM-DD_HHMMSS.xlsx
"""
import difflib
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    requests = None

# Column indices in run log (1-based): Source file=1, ..., OCR output=6, Invoice summary=7
RUN_LOG_HEADERS = ("Source file", "Start timestamp", "OCR time (s)", "Summary time (s)", "Who processed", "OCR output", "Invoice summary")
SOURCE_COL = 1
OCR_OUTPUT_COL = 6
INVOICE_SUMMARY_COL = 7
_EXCEL_CELL_MAX = 32000


def _trunc(s: str, max_len: int = _EXCEL_CELL_MAX) -> str:
    if not s or len(s) <= max_len:
        return s or ""
    return s[: max_len - 3] + "..."


def _diff_text(a: str, b: str, fromfile: str = "run1", tofile: str = "run2", max_lines: int = 500) -> str:
    """Produce a unified diff between two text strings. Truncate if too long."""
    a_lines = (a or "").splitlines(keepends=True)
    b_lines = (b or "").splitlines(keepends=True)
    lines = list(difflib.unified_diff(a_lines, b_lines, fromfile=fromfile, tofile=tofile, lineterm=""))
    if len(lines) > max_lines:
        lines = lines[:max_lines] + [f"\n... (diff truncated, {len(lines) - max_lines} more lines)\n"]
    return "".join(lines).strip() or "(no difference)"


def _same_percent(a: str, b: str) -> float | None:
    """Return similarity of a and b as a percentage (0–100). None if both empty."""
    a = (a or "").strip()
    b = (b or "").strip()
    if not a and not b:
        return None
    if not a or not b:
        return 0.0
    ratio = difflib.SequenceMatcher(None, a, b).ratio()
    return round(ratio * 100, 2)


# Max chars to send to Gemini per text (avoid token limits)
_GEMINI_TEXT_MAX = 12000

METADATA_CONSISTENCY_PROMPT = """You are comparing two summaries of the SAME document from two different runs (run 1 and run 2).
Do NOT compare character-by-character. Check LOGICAL CONSISTENCY of metadata between the two:
- Key amounts (totals, VAT, amounts due) match.
- Dates (invoice date, due date) match.
- Parties (vendor, client names) match.
- Invoice/reference numbers match if present.

Reply with:
- If consistent: one line: "Consistent"
- If inconsistent: first line "Inconsistent: <brief reason>", then a second line "Summary: <what was inconsistent>" with a short summary of which metadata differed (e.g. amounts, dates, parties).
- If one summary is missing: one line "N/A (only in one run)" """


def _gemini_metadata_consistent(summary1: str, summary2: str) -> tuple[str, str]:
    """
    Ask Gemini whether metadata in summary1 and summary2 (same document, two runs) is logically consistent.
    Returns (verdict_line, summary_of_inconsistent). summary_of_inconsistent is non-empty only when inconsistent.
    """
    if not requests:
        return ("(requests not installed)", "")
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not key:
        return ("(GEMINI_API_KEY not set)", "")
    s1 = (summary1 or "").strip()
    s2 = (summary2 or "").strip()
    if not s1 or not s2:
        return ("N/A (only in one run)", "")
    if len(s1) > _GEMINI_TEXT_MAX:
        s1 = s1[:_GEMINI_TEXT_MAX] + "\n... [truncated]"
    if len(s2) > _GEMINI_TEXT_MAX:
        s2 = s2[:_GEMINI_TEXT_MAX] + "\n... [truncated]"
    api_version = "v1beta" if os.environ.get("GEMINI_BETA", "").strip().lower() in ("1", "true", "yes") else "v1"
    model = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash").strip()
    url = f"https://generativelanguage.googleapis.com/{api_version}/models/{model}:generateContent"
    headers = {"x-goog-api-key": key, "Content-Type": "application/json"}
    user_content = (
        f"{METADATA_CONSISTENCY_PROMPT}\n\n"
        "--- Summary from run 1 ---\n" + s1 + "\n\n--- Summary from run 2 ---\n" + s2
    )
    payload = {
        "contents": [{"parts": [{"text": user_content}]}],
        "generationConfig": {"maxOutputTokens": 512, "temperature": 0.1},
    }
    try:
        r = requests.post(url, json=payload, headers=headers, timeout=60)
        r.raise_for_status()
        data = r.json()
        parts = (data.get("candidates", [{}])[0].get("content", {}).get("parts", []))
        text = (parts[0].get("text", "") if parts else "").strip()
        if not text:
            return ("(empty response)", "")
        lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
        verdict = lines[0] if lines else ""
        summary_inconsistent = ""
        for i, ln in enumerate(lines[1:], start=1):
            if ln.lower().startswith("summary:"):
                summary_inconsistent = ln[8:].strip()
                # Append any following lines
                for ln2 in lines[i + 1 :]:
                    summary_inconsistent += " " + ln2
                break
        return (verdict, summary_inconsistent.strip())
    except requests.exceptions.RequestException as e:
        return (f"(API error: {e})", "")
    except (KeyError, IndexError):
        return ("(unexpected response)", "")


def read_run_log_rows(xlsx_path: Path, from_file: str) -> list[dict]:
    """Read run log xlsx; return list of dicts with source_file, ocr_output, invoice_summary, from_file."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        source = (row[SOURCE_COL - 1] or "").strip()
        ocr = (row[OCR_OUTPUT_COL - 1] or "").strip() if len(row) >= OCR_OUTPUT_COL else ""
        summary = (row[INVOICE_SUMMARY_COL - 1] or "").strip() if len(row) >= INVOICE_SUMMARY_COL else ""
        rows.append({"source_file": source, "ocr_output": ocr, "invoice_summary": summary, "from_file": from_file})
    wb.close()
    return rows


def group_by_source_file(rows_from_all: list[dict], file_order: list[str]) -> dict[str, dict[str, dict]]:
    """
    Group rows by source_file. file_order = [file1_name, file2_name, ...].
    Returns: { source_file: { file1_name: {ocr, summary}, file2_name: {ocr, summary}, ... }, ... }
    """
    grouped: dict[str, dict[str, dict]] = {}
    for row in rows_from_all:
        sf = row["source_file"]
        if not sf:
            continue
        if sf not in grouped:
            grouped[sf] = {f: {"ocr": "", "summary": ""} for f in file_order}
        from_f = row["from_file"]
        if from_f in grouped[sf]:
            grouped[sf][from_f] = {"ocr": row["ocr_output"], "summary": row["invoice_summary"]}
    return grouped


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    argv = sys.argv[1:]

    # Parse --input / --output / --no-gemini
    input_paths: list[Path] = []
    output_path: Path | None = None
    use_gemini = True
    i = 0
    while i < len(argv):
        if argv[i] == "--input" and i + 1 < len(argv):
            input_paths.append(Path(argv[i + 1]))
            i += 2
            continue
        if argv[i] == "--output" and i + 1 < len(argv):
            output_path = Path(argv[i + 1])
            i += 2
            continue
        if argv[i] == "--no-gemini":
            use_gemini = False
            i += 1
            continue
        if not argv[i].startswith("-"):
            input_paths.append(Path(argv[i]))
        i += 1

    # Default: use two most recent run logs in log_files/
    if not input_paths:
        log_dir = script_dir / "log_files"
        if not log_dir.exists():
            print("No input xlsx given and log_files/ not found. Pass exactly 2 files: python compare_ocr_run_log.py file1.xlsx file2.xlsx")
            sys.exit(1)
        candidates = sorted(log_dir.glob("invoices_run_log_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if len(candidates) < 2:
            print("Need 2 run log files to compare. Pass them: python compare_ocr_run_log.py file1.xlsx file2.xlsx")
            sys.exit(1)
        input_paths = [candidates[0], candidates[1]]
        print(f"Using: {input_paths[0].name} vs {input_paths[1].name}")

    if len(input_paths) != 2:
        print("Exactly 2 run log files must be passed. Usage: python compare_ocr_run_log.py file1.xlsx file2.xlsx")
        sys.exit(1)

    # Default output path (with timestamp)
    if output_path is None:
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_path = script_dir / "log_files" / f"ocr_comparison_{stamp}.xlsx"
    if not output_path.is_absolute():
        output_path = script_dir / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if use_gemini and not os.environ.get("GEMINI_API_KEY", "").strip():
        use_gemini = False
        print("GEMINI_API_KEY not set; skipping Metadata consistent (Gemini) column. Use --no-gemini to hide this message.")

    # Resolve paths and build file order (use short names for headers)
    resolved: list[Path] = []
    file_order: list[str] = []
    for p in input_paths:
        if not p.is_absolute():
            p = script_dir / p
        if not p.exists():
            print(f"Skip (not found): {p}")
            continue
        resolved.append(p)
        file_order.append(p.name)

    if len(resolved) != 2:
        print("Exactly 2 existing run log files are required.")
        sys.exit(1)

    # Collect rows from each file (tagged with from_file)
    all_rows: list[dict] = []
    for p in resolved:
        try:
            rows = read_run_log_rows(p, from_file=p.name)
            all_rows.extend(rows)
            print(f"Read {len(rows)} rows from {p.name}")
        except Exception as e:
            print(f"Error reading {p}: {e}", file=sys.stderr)

    if not all_rows:
        print("No rows to compare.")
        sys.exit(1)

    # Group by source file; each source has { file1: {ocr, summary}, file2: {ocr, summary}, ... }
    grouped = group_by_source_file(all_rows, file_order)
    f1_name = file_order[0]
    f2_name = file_order[1]

    # Build comparison: one row per source file; columns = Source | OCR(f1) | Summary(f1) | OCR(f2) | Summary(f2) | Difference | Same % | [Gemini]
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    headers = ["Source file"]
    for fn in file_order:
        headers.append(f"OCR ({fn})")
        headers.append(f"Summary ({fn})")
    headers.extend([
        "Difference (OCR run1 vs run2)",
        "Same % (OCR)",
        "Difference (Summary run1 vs run2)",
        "Same % (Summary)",
        "Metadata consistent (Gemini)" if use_gemini else None,
        "Summary of what was inconsistent" if use_gemini else None,
    ])
    headers = [h for h in headers if h is not None]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    col_source = 1
    n_files = len(file_order)
    # Two columns per file: OCR then Summary. Then OCR diff, OCR same%, Summary diff, Summary same%, Gemini, Summary inconsistent.
    col_diff_ocr = 2 + n_files * 2
    col_same_ocr = col_diff_ocr + 1
    col_diff_summary = col_same_ocr + 1
    col_same_summary = col_diff_summary + 1
    col_gemini = col_same_summary + 1
    col_summary_inconsistent = col_gemini + 1

    for r, (source_file, by_file) in enumerate(sorted(grouped.items()), start=2):
        ws.cell(row=r, column=col_source, value=source_file)
        for i, fn in enumerate(file_order):
            data = by_file.get(fn, {"ocr": "", "summary": ""})
            ocr = data["ocr"] or ""
            summary = data["summary"] or ""
            ws.cell(row=r, column=2 + i * 2, value=_trunc(ocr))
            ws.cell(row=r, column=3 + i * 2, value=_trunc(summary))

        ocr1 = (by_file.get(f1_name) or {}).get("ocr") or ""
        ocr2 = (by_file.get(f2_name) or {}).get("ocr") or ""
        summary1 = (by_file.get(f1_name) or {}).get("summary") or ""
        summary2 = (by_file.get(f2_name) or {}).get("summary") or ""

        # OCR comparison
        if ocr1 and ocr2:
            diff_ocr = _diff_text(ocr1, ocr2, fromfile=f1_name, tofile=f2_name)
            same_pct_ocr = _same_percent(ocr1, ocr2)
        else:
            diff_ocr = "N/A (only in one run)"
            same_pct_ocr = None

        # Summary comparison (diff + percentage)
        if summary1 and summary2:
            diff_summary = _diff_text(summary1, summary2, fromfile=f1_name, tofile=f2_name)
            same_pct_summary = _same_percent(summary1, summary2)
        else:
            diff_summary = "N/A (only in one run)"
            same_pct_summary = None

        ws.cell(row=r, column=col_diff_ocr, value=_trunc(diff_ocr))
        ws.cell(row=r, column=col_same_ocr, value=same_pct_ocr if same_pct_ocr is not None else "")
        ws.cell(row=r, column=col_diff_summary, value=_trunc(diff_summary))
        ws.cell(row=r, column=col_same_summary, value=same_pct_summary if same_pct_summary is not None else "")

        if use_gemini:
            if summary1 and summary2:
                meta, summary_inconsistent = _gemini_metadata_consistent(summary1, summary2)
            else:
                meta, summary_inconsistent = "N/A (only in one run)", ""
            ws.cell(row=r, column=col_gemini, value=_trunc(meta, max_len=1000))
            ws.cell(row=r, column=col_summary_inconsistent, value=_trunc(summary_inconsistent, max_len=2000))
            if (r - 1) % 5 == 0 or r == 2:
                preview = meta[:80] + "..." if len(meta) > 80 else meta
                print(f"  Gemini check {source_file}: {preview}")

    wb.save(output_path)
    print(f"Wrote {len(grouped)} rows (one per source file) to {output_path}")


if __name__ == "__main__":
    main()
